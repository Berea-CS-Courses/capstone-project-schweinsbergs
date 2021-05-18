import pandas as pd  # for dataframe
import pickle  # to save dataframe
from recipe_scrapers import scrape_me  # recipe scraper
import openpyxl  # for working with excel files
import tkinter as tk # for user interface
from tkinter import simpledialog # for user interface
from pandasgui import show # for showing the dataframe



# Assistance with the pickle code from:
# https://www.youtube.com/watch?v=WkIW0YLoQEE&list=PLQVvvaa0QuDc-3szzjeP6N6b0aDrrKyL-&t=507s&ab_channel=sentdex

# Assistance with sitemap scraping from:
# https://github.com/dsottimano/xmlsitemap-extractor-google-sheets/blob/master/xmlscript.js
#######################################################################################################################


class testScrape:
    """
    The class for scraping and storing recipes into a dataframe, and then pickling them.
    """

    def __init__(self, scraper, recipeInfo, df, title, yields, ingredients, recipeLoad, user_ingredients,
                 instructions, link):
        """
        self.scraper = the web scraper
        :param scraper:
        self.recipeInfo = data pulled from scraper for data frame
        self.df = data frame
        self.yields = the yields of the recipe
        self.ingredients = the ingredients of the recipe
        self.recipeLoad = loaded pickle file contents
        self.user_ingredients = the user's inputted ingredients
        self.instructions = the user's inputted instructions
        """

        self.scraper = scraper
        self.recipeInfo = recipeInfo
        self.df = df
        self.title = title
        self.yields = yields
        self.ingredients = ingredients
        self.recipeLoad = recipeLoad
        self.user_ingredients = user_ingredients
        self.instructions = instructions
        self.url = link

    def dataframe(self):
        """
        This sets up the dataframe. It opens up the excel sheet that contains all of the recipes URLS, reads it cell
        by cell, and populates the dataframe with each recipe.
        :return:
        """

        # Assigns each column of the data frame with a value.
        self.recipeInfo = {'Title': [self.title],
                           'yields': [self.yields],
                           'ingredients': [self.ingredients],
                           'Instructions': [self.instructions],
                           'Link': [self.url]}

        # Actually creates the dataframe based on those values.

        self.df = pd.DataFrame(self.recipeInfo,
                               columns=['Title', 'yields', 'ingredients', 'Instructions', 'Link'])
        # Now we have an empty dataframe with only titles, so we're going to open up the excel sheet. readurls is what
        # we use to actually READ the sheets.
        excellsheet = openpyxl.load_workbook("xmlinfo.xlsx")
        readurls = excellsheet.active

        # This iterates down the excel sheet and feeds urls to the web scraper. The web scraper then scrapes that url,
        # and places the data into the dataframe.
        for i in range(1, readurls.max_row + 1):  # changing the number @ range will change where the program reads
            recipeurls = readurls.cell(row=i, column=1)

            self.scraper = scrape_me(recipeurls.value)
            self.title = self.scraper.title()
            self.yields = self.scraper.yields()
            self.ingredients = self.scraper.ingredients()
            self.instructions = self.scraper.instructions()
            self.url = recipeurls.value

            addition = {'Title': self.title, 'yields': self.yields,
                        'ingredients': self.ingredients, 'Instructions': self.instructions, 'Link': self.url}

            self.df = self.df.append(addition, ignore_index=True)
            print("Succeeded scrape at row", recipeurls)  # feedback so I'm not sitting in idle hell

        print(self.df)

    def pickle_jar(self):
        """
        Function to "pickle" my dataframe, which saves it as a pickle file.
        :return:
        """

        pickle_out = open('pickleRecipe.pickle', 'wb')  # Creates pickleRecipe.pickle file, wb = write bytes
        pickle.dump(self.df, pickle_out)  # dumps dataframe into pickle_out, the pickle file
        pickle_out.close()

    def open_pickle_jar(self):
        """
        Opens the pickle! This lets the program read from the pickle file so that the data can be used later.
        :return:
        """

        pickle_in = open('pickleRecipe.pickle', 'rb')  # opens the pickle, now rb = read bytes
        self.recipeLoad = pickle.load(pickle_in)  # reads pickle file

    def compare_ingredients(self):
        """
        Compares the dataframe's ingredients list and the user's.
        help from: https://towardsdatascience.com/dealing-with-list-values-in-pandas-dataframes-a177e534f173

        help from: https://stackoverflow.com/questions/35240528/reverse-dataframes-rows-order-with-pandas

        So much thanks to Mario and Jesse! Emely and Kaleb helped, too. :)
        Matches help from Ezra and Conner! Thank you!

        fractionlist is an empty list that holds our percentages later
        ingredientlist splits up our dataframe cells into a list
        matches is an empty list that holds our matches
        specficmatches is the user's matches
        :return:
        """
        # For each cell in the ingredients column, and each word in the list of ingredients, split them up into
        # a list. It's then compared against the user input list-- The break means the user list will stop searching
        # for that word in the ingredients. Prevents multiple matches for one ingredient.
        fractionList = []
        #specificmatchesarray = []
        for ind in self.recipeLoad.index:
           ingredientList = self.recipeLoad['ingredients'][ind]
           matches = []
           specficmatches = []
           for userWord in self.user_ingredients:
               matched = False
               for ingredient in ingredientList:
                   if userWord.lower() in ingredient.lower():
                    matched = True
                    specficmatches.append(userWord.lower())
                    break
               matches.append(matched)
               #specificmatchesarray.append(specficmatches)
               #print(specificmatchesarray)




           # Calculate the percentage matched here by calculating the length, put that on the dataframe, and reverse it.
           if matches.count(True) == 0:
               fraction = 0
           else:
            fraction = (matches.count(True) / len(ingredientList)) * 100
           if fraction >= 100:
               fraction = 100
           fractionList.append(fraction)
           #print(self.recipeLoad['Title'][ind],fraction)
        self.recipeLoad['Percentage'] = fractionList
        #self.recipeLoad['Matches'] = specificmatchesarray
        self.recipeLoad.sort_values(by='Percentage', inplace=True)
        reversed_dataframe = self.recipeLoad.iloc[::-1]
        show(reversed_dataframe)



    def user_input_window(self):
        """
        Help from: https://djangocentral.com/creating-user-input-dialog/
        :return:
        """

        canvas = tk.Tk()
        canvas.withdraw()
        self.entry = simpledialog.askstring(title="Ingredient Input",
                                  prompt="Please enter your ingredients separated by a space!")

        self.user_ingredients = self.entry.split()
        #print(self.user_ingredients)




def main():
    """
    Currently calls everything so I can test it.
    :return:open('pickleRecipe.pickle', 'wb')
    """
    test = testScrape('scraper', 'recipeInfo', 'df', 'title', 'yields', 'ingredients', 'recipeLoad',
                      'user_ingredients', 'Instructions', 'Link')

    #test.dataframe()
    #test.pickle_jar()
    test.open_pickle_jar()
    test.user_input_window()
    test.compare_ingredients()


main()